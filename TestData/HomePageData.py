import openpyxl

class HomePageData:

    test_HomePage_data =[{"firstname":"Lo", "email": "lo1@gmail.com", "gender": "Male"}, {"firstname":"Remy", "lastname": "remy2@gmail.com", "gender": "Female"} ]

    @staticmethod
    def getTestData(test_case_name):

        book = openpyxl.load_workbook("C:\\Users\\crab\\PycharmProjects\\PythonSelfFramework\\TestData\\pythonDemo.xlsx")
        sheet = book.active
        Dict = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]