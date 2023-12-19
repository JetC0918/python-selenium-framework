import openpyxl

book = openpyxl.load_workbook("C:\\Users\\crab\\PycharmProjects\\PythonSelfFramework\\TestData\\pythonDemo.xlsx")
sheet = book.active
Dict = {}

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        Dict[sheet.cell(row=1,column=j).value]= sheet.cell(row=i,column=j).value

print (Dict)